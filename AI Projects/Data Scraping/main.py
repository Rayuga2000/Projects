import pandas as pd
import os
import re
import pathlib
import textwrap
import google.generativeai as genai
from IPython.display import display
from IPython.display import Markdown
#from openpyxl import Workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from PIL import Image
from openai import OpenAI
import requests
import cv2
import numpy as np
import matplotlib.pyplot as plt
import sys

## gemini functions start ##

def to_markdown(text):
  text = text.replace('•', '  *')
  return Markdown(textwrap.indent(text, '> ', predicate=lambda _: True))

def modelList_gemini():
    for m in genai.list_models():
        if 'generateContent' in m.supported_generation_methods:
            print(m.name)

def gemini_response(usrInput):
  model = genai.GenerativeModel('models/gemini-1.5-pro-latest')

  response = model.generate_content("detailed analysis about "+usrInput+" tourist attractions following this format: * **({Country Only}) ({State Only}) ({City Only}) ({Attraction Name Only}):** {Place Description} (must start with a capital letter and Description should be 2500 characters long).")
  #print(response.text)
  # to_markdown(response.text)

  ## Scrap data
  pattern = r"(?<=\*)\(([^)]+)\)\s+\(([^)]+)\)\s+\(([^)]+)\)\s+\(([^)]+)\):\*\*\s(.*)"

  return re.findall(pattern, response.text, re.MULTILINE)

def gemini_description(usrInput):
    global  data, gemini_places

    while gemini_places==0:
      print('API Called')
      x=gemini_response(usrInput)
      gemini_places = len(x)
      # print('Error occured')
      # sys.exit(1)

    print(f'Found {gemini_places} places')
    for i in x:
        #print(i)
        data['Country'].append(i[0])
        data['State'].append(i[1])
        data['City'].append(i[2])
        data['Attraction Name'].append(i[3])
        data['Gemini Description'].append(i[4])

## gemini functions end ##

def gen_Image(attraction):
  global data

  image_response = client.images.generate(
    model="dall-e-3",
    prompt=attraction,
    n=1,
    size="1024x1024"
  )

  img_url = image_response.data[0].url
  data['Images'].append(img_url)
  #print(img_url)

  # with open(f'{attraction}.png','wb') as file:
    # file.write(requests.get(img_url).content)

def tweak_xl():
  if os.path.exists('final.xlsx'):
    wb = openpyxl.load_workbook(filename="final.xlsx")
    ws = wb.active

    df = pd.read_excel('temp.xlsx', skiprows=1)
    #df = df[1:]
    #print(df)

    for row in dataframe_to_rows(df,index=False):
      #print(row)
      ws.append(row)
  else:
    wb = openpyxl.load_workbook(filename="temp.xlsx")
    ws = wb.active

  ws.column_dimensions['A'].width = 20
  ws.column_dimensions['B'].width = 20
  ws.column_dimensions['C'].width = 20
  ws.column_dimensions['D'].width = 20
  ws.column_dimensions['E'].width = 100
  ws.column_dimensions['F'].width = 10

  for col in ws.iter_cols(min_row=2, min_col=1, max_col=6):
    #print(col)
    for cell in col:
      #print(cell.col_idx)
      if cell.col_idx==6:
        cell.hyperlink = cell.value #str(cell.value)
        cell.value = "Click Here"

      cell.alignment=openpyxl.styles.Alignment(horizontal='justified',vertical='center')

  wb.save('final.xlsx')
  os.remove('temp.xlsx')
  print('Excel file saved as final.xlsx')

if __name__=='__main__':
    GOOGLE_API_KEY = 'AIzaSyBaZbMC10O4ovbY_fac-slH5ByjQR3_WkE'
    genai.configure(api_key=GOOGLE_API_KEY)

    OPENAI_API_KEY = 'sk-proj-S6NBF2oFTtcU193xuEDLT3BlbkFJYtIKYTHqoeueE8ad7ENn'
    client = OpenAI(api_key=OPENAI_API_KEY, )

    while(True):
      gemini_places=0

      # dictionary for saving data to excel
      data={'Country': [], 'State': [], 'City': [], 'Attraction Name': [], 'Gemini Description': [], 'Images': []}

      usrInput=input('\nEnter name of a Country/State/City (to exit enter <ctrl+c>): ')

      print('Getting Data from Google...')
      gemini_description(usrInput)

      print('Generating Images...')
      for i in data['Attraction Name']:
        gen_Image(i)

      df = pd.DataFrame(data)
      # df.to_csv('my_data.csv', index=False)
      df.to_excel('temp.xlsx', index=False)

      print('Saving Output...')
      tweak_xl()
    